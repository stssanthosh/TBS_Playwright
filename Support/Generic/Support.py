import os
from openpyxl import load_workbook
from configparser import ConfigParser



def get_config():
    config = ConfigParser()
    current_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.abspath(os.path.join(current_dir, "..", "..", "config.ini"))
    config.read(config_path)  # or full path if needed
    return config["Environment"]

def file_sheet_index(file_name, sheet_name):
    """
    Reads an Excel sheet, creates environment variables dynamically for each cell
    Format: {SheetName}{ModifiedColumnName}{RowNumber} = CellValue
    """

    project_directory = os.environ["PROJECT_DIRECTORY"]

    file_path = os.path.join(project_directory, f"TestData/{file_name}.xlsx")
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    max_column = ws.max_column
    max_row = ws.max_row

    # Read column headers (row 1)
    columns = []
    for col_idx in range(1, max_column + 1):
        column_header = ws.cell(row=1, column=col_idx).value
        columns.append(column_header)

    # Iterate columns and rows to build env variables
    for col_idx, column_name in enumerate(columns, start=1):
        modified_column = column_header_modify(column_name)
        for row_idx in range(2, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            env_key = f"{sheet_name}{modified_column}{row_idx}"

            os.environ[env_key] = str(value) if value is not None else ""

    return os.environ, max_row  # Optional: return for further use

def read_excel(sheet, row, column_header):
    """
    Retrieves a value from environment variable based on sheet, column header, and row.
    Format: {Sheet}{ModifiedColumn}{Row}
    """
    # Apply your column header modify logic
    modified_column = column_header_modify(column_header)

    # Build the env variable key
    env_key = f"{sheet}{modified_column}{row}"

    # Fetch value from environment
    value = os.getenv(env_key)
    return value

def column_header_modify(column):
    """
    Custom logic: Remove spaces, convert to uppercase.
    """
    column = column.replace(" ", "")
    column = column.upper()
    return column
